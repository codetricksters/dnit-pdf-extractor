import io
from typing import List

import pandas as pd
from fastapi import APIRouter, File, Request, UploadFile
from fastapi.responses import StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..services.extractor import EXPECTED_COLUMNS, extract_from_pdf

router = APIRouter()

# Resolved at import time from the app package so the path is always correct
# regardless of the working directory the server is started from.
from pathlib import Path
templates = Jinja2Templates(directory=str(Path(__file__).parent.parent / "templates"))


@router.get("/", include_in_schema=False)
async def index(request: Request):
    return templates.TemplateResponse(request, "index.html")


@router.post("/upload", tags=["extraction"])
async def upload(files: List[UploadFile] = File(...)):
    all_rows: list[dict] = []
    for file in files:
        content = await file.read()
        rows = extract_from_pdf(content, file.filename or "unknown.pdf")
        all_rows.extend(rows)

    df = pd.DataFrame(all_rows, columns=EXPECTED_COLUMNS + ["Source_File"])

    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(df.columns.tolist())
    for row in df.itertuples(index=False):
        ws.append(list(row))

    for col_cells in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=medicao.xlsx"},
    )
