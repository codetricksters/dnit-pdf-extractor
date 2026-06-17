"""Tests for POST /upload — file upload and Excel generation."""
import io
from unittest.mock import patch

import openpyxl
import pytest

from .conftest import make_rows
from .pdf_factory import corrupt_pdf, no_table_pdf


def _post_files(client, files_bytes: dict[str, bytes]):
    files = [
        ("files", (name, data, "application/pdf"))
        for name, data in files_bytes.items()
    ]
    return client.post("/upload", files=files)


def _post_mocked(client, filenames: list[str]):
    """Post PDFs where extract_from_pdf is mocked to return sample rows."""
    def fake_extract(file_bytes, source_name):
        return make_rows(source_name)

    with patch("app.routers.upload.extract_from_pdf", side_effect=fake_extract):
        files = [
            ("files", (name, b"fakepdfbytes", "application/pdf"))
            for name in filenames
        ]
        return client.post("/upload", files=files)


# --- Input validation ---

def test_upload_no_files_returns_422(client):
    response = client.post("/upload")
    assert response.status_code == 422


def test_upload_corrupt_pdf_returns_422(client):
    response = _post_files(client, {"corrupt.pdf": corrupt_pdf()})
    assert response.status_code == 422
    assert "corrupt.pdf" in response.json()["detail"]


def test_upload_no_table_pdf_returns_422(client):
    response = _post_files(client, {"empty.pdf": no_table_pdf()})
    assert response.status_code == 422


# --- Happy path (mocked extractor) ---

def test_upload_valid_pdf_returns_200(client):
    response = _post_mocked(client, ["medicao.pdf"])
    assert response.status_code == 200


def test_upload_response_content_type_is_xlsx(client):
    response = _post_mocked(client, ["medicao.pdf"])
    ct = response.headers["content-type"]
    assert "spreadsheetml" in ct or "openxmlformats" in ct


def test_upload_response_has_attachment_header(client):
    response = _post_mocked(client, ["medicao.pdf"])
    disposition = response.headers.get("content-disposition", "")
    assert "attachment" in disposition
    assert ".xlsx" in disposition


def test_upload_excel_has_single_sheet(client):
    response = _post_mocked(client, ["medicao.pdf"])
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert len(wb.sheetnames) == 1
    assert wb.sheetnames[0] == "Sheet1"


def test_upload_excel_has_header_row(client):
    response = _post_mocked(client, ["medicao.pdf"])
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    headers = [cell.value for cell in wb.active[1]]
    assert "Source_File" in headers
    assert "Serviço" in headers


def test_upload_excel_data_rows_present(client):
    response = _post_mocked(client, ["medicao.pdf"])
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    data_rows = list(wb.active.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 2


def test_upload_source_file_column_matches_filename(client):
    response = _post_mocked(client, ["medicao.pdf"])
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    source_col_idx = headers.index("Source_File") + 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        assert row[source_col_idx - 1] == "medicao.pdf"


def test_upload_multiple_files_merged_into_one_sheet(client):
    response = _post_mocked(client, ["file1.pdf", "file2.pdf"])
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert len(wb.sheetnames) == 1
    data_rows = list(wb.active.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 4  # 2 rows per file


def test_upload_multiple_files_source_column_tracks_origin(client):
    response = _post_mocked(client, ["alpha.pdf", "beta.pdf"])
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    source_col_idx = headers.index("Source_File") + 1
    sources = {row[source_col_idx - 1] for row in ws.iter_rows(min_row=2, values_only=True)}
    assert "alpha.pdf" in sources
    assert "beta.pdf" in sources
